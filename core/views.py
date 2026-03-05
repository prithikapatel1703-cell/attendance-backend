from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework import status
from django.contrib.auth import login
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from io import BytesIO
from openpyxl import load_workbook, Workbook
import datetime
from django.db.models import Sum
from .models import User, Attendance, Department, Subject
from .serializers import (
    RegisterSerializer, LoginSerializer, AttendanceSerializer, UserSerializer,
    DepartmentSerializer, SubjectSerializer,
)


@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):

    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.validated_data['user']
        login(request, user)

        role = user.role
        if user.is_superuser and request.data.get("role") == "admin":
            role = "admin"

        return Response({
            "message": "Login successful",
            "id": user.id,
            "email": user.email,
            "role": role,
            "username": user.username,
            "full_name": user.full_name or user.username,
            "department": user.department or "",
        })

    return Response(serializer.errors, status=400)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def attendance_view(request):

    if request.method == 'GET':
        if request.user.role == 'student':
            records = Attendance.objects.filter(student=request.user)
        else:
            records = Attendance.objects.all()

        # Optional date range filtering: from_date, to_date in YYYY-MM-DD format
        from_date_str = request.query_params.get('from_date') or request.GET.get('from_date')
        to_date_str = request.query_params.get('to_date') or request.GET.get('to_date')
        from_date = to_date = None
        try:
            if from_date_str:
                from_date = datetime.date.fromisoformat(from_date_str)
            if to_date_str:
                to_date = datetime.date.fromisoformat(to_date_str)
        except (TypeError, ValueError):
            # Ignore invalid filters and return full data instead of erroring out
            from_date = to_date = None

        if from_date:
            records = records.filter(date__gte=from_date)
        if to_date:
            records = records.filter(date__lte=to_date)

        total_classes = records.count()
        present_count = records.filter(status__iexact='present').count()

        # Prefer hours-based percentage when available (attended_hours / total_hours)
        hours_agg = records.aggregate(attended=Sum('attended_hours'), total=Sum('total_hours'))
        attended_hours_sum = hours_agg['attended'] or 0
        total_hours_sum = hours_agg['total'] or 0
        percentage = 0
        if total_hours_sum and total_hours_sum > 0:
            percentage = (attended_hours_sum / total_hours_sum) * 100
        elif total_classes > 0:
            percentage = (present_count / total_classes) * 100

        serializer = AttendanceSerializer(records, many=True)

        return Response({
            "total_classes": total_classes,
            "present_count": present_count,
            "attended_hours": attended_hours_sum,
            "total_hours": total_hours_sum,
            "attendance_percentage": round(percentage, 2),
            "records": serializer.data
        })

    elif request.method == 'POST':
        if request.user.role not in ['faculty', 'admin']:
            return Response({"error": "Not authorized"}, status=403)

        data = request.data
        if isinstance(data, list):
            def parse_date_for_post(value):
                if value is None:
                    return None
                if isinstance(value, datetime.date):
                    return value if not isinstance(value, datetime.datetime) else value.date()
                if isinstance(value, datetime.datetime):
                    return value.date()
                try:
                    return datetime.date.fromisoformat(str(value).strip())
                except (TypeError, ValueError):
                    return None

            saved_count = 0
            errors = []
            for i, item in enumerate(data):
                if not isinstance(item, dict):
                    errors.append({"index": i, "errors": {"detail": "Invalid item, expected object."}})
                    continue
                try:
                    student_id = int(item.get('student'))
                except (TypeError, ValueError):
                    student_id = None
                subject = item.get('subject')
                date_val = item.get('date')
                status_val = item.get('status')
                if student_id is None or not str(subject).strip():
                    errors.append({"index": i, "errors": {"detail": "Missing student or subject."}})
                    continue
                if not str(status_val).strip().lower() in ('present', 'absent'):
                    errors.append({"index": i, "errors": {"detail": "Status must be 'present' or 'absent'."}})
                    continue
                date_obj = parse_date_for_post(date_val)
                if not date_obj:
                    errors.append({"index": i, "errors": {"detail": f"Invalid date: {date_val}"}})
                    continue
                if not User.objects.filter(pk=student_id, role='student').exists():
                    errors.append({"index": i, "errors": {"detail": f"Student id {student_id} not found."}})
                    continue
                status_normalized = str(status_val).strip().lower()
                if status_normalized in ('present', 'p'):
                    status_normalized = 'present'
                elif status_normalized in ('absent', 'a'):
                    status_normalized = 'absent'
                _, created = Attendance.objects.update_or_create(
                    student_id=student_id,
                    subject=str(subject).strip(),
                    date=date_obj,
                    defaults={"status": status_normalized},
                )
                saved_count += 1
            return Response({
                "created": saved_count,
                "records": [],  # frontend refetches via GET
                "errors": errors if errors else None,
            }, status=201 if saved_count > 0 else (400 if errors else 201))

        serializer = AttendanceSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_list_view(request):
    role = request.query_params.get('role', 'student')
    is_admin = request.user.role == 'admin' or request.user.is_superuser
    is_faculty = request.user.role == 'faculty'

    if not is_admin and not is_faculty:
        return Response({"detail": "Not allowed to list users."}, status=403)
    if is_faculty and role != 'student':
        return Response({"detail": "Faculty can only list students."}, status=403)

    qs = User.objects.filter(role=role).order_by('id')
    if is_faculty:
        if request.user.department:
            qs = qs.filter(department=request.user.department)
            section = request.query_params.get('section', '').strip()
            if section:
                qs = qs.filter(section=section)
            year = request.query_params.get('year', '').strip()
            if year:
                qs = qs.filter(year=year)
        else:
            qs = User.objects.none()

    serializer = UserSerializer(qs, many=True)
    return Response(serializer.data)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def user_detail_view(request, pk):
    try:
        target = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({"detail": "User not found."}, status=404)

    is_admin = request.user.role == 'admin' or request.user.is_superuser
    is_faculty = request.user.role == 'faculty'
    is_self = request.user.id == target.id
    faculty_can_edit_student = is_faculty and target.role == 'student' and target.department == request.user.department

    if not (is_admin or is_self or faculty_can_edit_student):
        return Response({"detail": "Not allowed to access this user."}, status=403)

    if request.method == 'GET':
        serializer = UserSerializer(target)
        return Response(serializer.data)

    elif request.method == 'PATCH':
        serializer = UserSerializer(target, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)

    elif request.method == 'DELETE':
        if not is_admin:
            return Response({"detail": "Only admin can delete users."}, status=403)
        if is_self:
            return Response({"detail": "You cannot delete your own account."}, status=400)
        if target.is_superuser:
            return Response({"detail": "Cannot delete superuser."}, status=400)
        target.delete()
        return Response(status=204)


# --- Departments (Branches) - Admin only ---
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def department_list_view(request):
    if request.user.role != 'admin' and not request.user.is_superuser:
        return Response({"detail": "Admin only."}, status=403)
    if request.method == 'GET':
        qs = Department.objects.all().order_by('code')
        return Response(DepartmentSerializer(qs, many=True).data)
    # POST
    serializer = DepartmentSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def department_detail_view(request, pk):
    if request.user.role != 'admin' and not request.user.is_superuser:
        return Response({"detail": "Admin only."}, status=403)
    try:
        dept = Department.objects.get(pk=pk)
    except Department.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if request.method == 'GET':
        return Response(DepartmentSerializer(dept).data)
    if request.method == 'PATCH':
        serializer = DepartmentSerializer(dept, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    if request.method == 'DELETE':
        dept.delete()
        return Response(status=204)


# --- Subjects: GET allowed for admin + faculty (so faculty portal can load assigned subjects), POST admin only ---
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def subject_list_view(request):
    is_admin = request.user.role == 'admin' or request.user.is_superuser
    is_faculty = request.user.role == 'faculty'
    if request.method == 'POST' and not is_admin:
        return Response({"detail": "Admin only."}, status=403)
    if request.method == 'GET' and not is_admin and not is_faculty:
        return Response({"detail": "Not allowed."}, status=403)
    if request.method == 'GET':
        qs = Subject.objects.select_related('department').all().order_by('department__code', 'year', 'semester', 'code')
        department = request.query_params.get('department', '').strip()
        if department:
            qs = qs.filter(department__code=department)
        year = request.query_params.get('year', '').strip()
        if year:
            qs = qs.filter(year=year)
        semester = request.query_params.get('semester', '').strip()
        if semester:
            qs = qs.filter(semester=semester)
        return Response(SubjectSerializer(qs, many=True).data)
    # POST: require department (id or code), year and semester optional (default '1')
    data = request.data.copy()
    dept_id = data.get('department')
    if not dept_id:
        return Response({"department": ["This field is required."]}, status=400)
    dept = Department.objects.filter(pk=dept_id).first() or Department.objects.filter(code=dept_id).first()
    if not dept:
        return Response({"department": ["Department not found."]}, status=400)
    data['department'] = dept.id
    if data.get('year') in (None, ''):
        data['year'] = '1'
    if data.get('semester') in (None, ''):
        data['semester'] = '1'
    serializer = SubjectSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def subject_detail_view(request, pk):
    if request.user.role != 'admin' and not request.user.is_superuser:
        return Response({"detail": "Admin only."}, status=403)
    try:
        subj = Subject.objects.get(pk=pk)
    except Subject.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if request.method == 'GET':
        return Response(SubjectSerializer(subj).data)
    if request.method == 'PATCH':
        serializer = SubjectSerializer(subj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    if request.method == 'DELETE':
        subj.delete()
        return Response(status=204)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_student_upload_view(request):
    is_admin = request.user.role == 'admin' or request.user.is_superuser
    if not is_admin:
        return Response({"detail": "Admin only."}, status=403)

    upload = request.FILES.get('file')
    if not upload:
        return Response({"detail": "No file uploaded. Use form field 'file'."}, status=400)
    if not str(upload.name).lower().endswith('.xlsx'):
        return Response({"detail": "Invalid file type. Please upload an .xlsx file."}, status=400)

    try:
        wb = load_workbook(filename=upload, data_only=True)
    except Exception:
        return Response({"detail": "Could not read Excel file. Make sure it is a valid .xlsx file."}, status=400)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return Response({"detail": "Excel file is empty."}, status=400)

    header_row = rows[0]
    headers = [str(v).strip().lower() if v is not None else '' for v in header_row]
    required_cols = ['full_name', 'roll_number', 'email', 'department', 'section', 'year']
    missing = [c for c in required_cols if c not in headers]
    if missing:
        return Response(
            {
                "detail": "Missing required columns in header row.",
                "missing_columns": missing,
                "expected_columns": required_cols,
                "received_headers": headers,
            },
            status=400,
        )

    idx = {name: headers.index(name) for name in required_cols}
    if 'password' in headers:
        idx['password'] = headers.index('password')

    created_count = 0
    skipped_existing = 0
    skipped_invalid = 0
    error_rows = []

    for row_number, row in enumerate(rows[1:], start=2):
        if row is None:
            continue
        if all((cell is None or str(cell).strip() == '') for cell in row):
            continue

        def _get(col_name):
            i = idx[col_name]
            if i >= len(row):
                return ''
            value = row[i]
            return '' if value is None else str(value).strip()

        full_name = _get('full_name')
        roll_number = _get('roll_number')
        email = _get('email')
        department = _get('department')
        section = _get('section')
        year = _get('year')

        if not roll_number or not email:
            skipped_invalid += 1
            error_rows.append({"row": row_number, "reason": "Missing roll_number or email."})
            continue

        if User.objects.filter(roll_number=roll_number, role='student').exists():
            skipped_existing += 1
            continue

        username_base = roll_number or email.split('@')[0]
        username = username_base
        suffix = 1
        while User.objects.filter(username=username).exists():
            username = f"{username_base}_{suffix}"
            suffix += 1

        password_raw = _get('password') if 'password' in idx else ''
        password = password_raw if password_raw else roll_number

        try:
            User.objects.create_user(
                username=username,
                email=email,
                password=password,
                role='student',
                full_name=full_name or username,
                roll_number=roll_number,
                department=department,
                section=section,
                year=str(year) if year is not None else '',
            )
        except Exception as exc:
            skipped_invalid += 1
            error_rows.append({"row": row_number, "reason": f"Failed to create user: {exc.__class__.__name__}"})
            continue

        created_count += 1

    return Response(
        {
            "created": created_count,
            "skipped_existing": skipped_existing,
            "skipped_invalid": skipped_invalid,
            "total_processed_rows": len(rows) - 1,
            "errors": error_rows,
            "note": "Password is taken from the 'password' column in Excel if present; otherwise roll_number is used.",
        }
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_attendance_excel_view(request):
    is_admin = request.user.role == 'admin' or request.user.is_superuser
    if not is_admin:
        return Response({"detail": "Admin only."}, status=403)

    wb = Workbook()

    # Students sheet
    ws_students = wb.active
    ws_students.title = "Students"
    ws_students.append(
        ["Roll Number", "Name", "Email", "Department", "Section", "Year", "Phone"]
    )
    for u in User.objects.filter(role="student").order_by("id"):
        ws_students.append(
            [
                u.roll_number or "",
                u.full_name or u.username,
                u.email or "",
                u.department or "",
                u.section or "",
                u.year or "",
                u.phone or "",
            ]
        )

    # Faculty sheet
    ws_faculty = wb.create_sheet("Faculty")
    ws_faculty.append(["Name", "Email", "Department", "Phone"])
    for u in User.objects.filter(role="faculty").order_by("id"):
        ws_faculty.append(
            [
                u.full_name or u.username,
                u.email or "",
                u.department or "",
                u.phone or "",
            ]
        )

    # Admins sheet
    ws_admins = wb.create_sheet("Admins")
    ws_admins.append(["Name", "Email"])
    for u in User.objects.filter(role="admin").order_by("id"):
        ws_admins.append([u.full_name or u.username, u.email or ""])

    # Departments sheet
    ws_depts = wb.create_sheet("Departments")
    ws_depts.append(["Code", "Name"])
    for d in Department.objects.all().order_by("code"):
        ws_depts.append([d.code, d.name])

    # Subjects sheet
    ws_subjects = wb.create_sheet("Subjects")
    ws_subjects.append(["Code", "Name", "Department Code", "Year", "Semester"])
    for s in Subject.objects.select_related("department").all().order_by(
        "department__code", "year", "semester", "code"
    ):
        ws_subjects.append(
            [
                s.code,
                s.name,
                s.department.code if s.department_id else "",
                s.year,
                s.semester,
            ]
        )

    # Attendance sheet
    ws_att = wb.create_sheet("Attendance")
    ws_att.append(["Date", "Student Roll Number", "Subject", "Status", "Attended Hours", "Total Hours"])
    student_by_id = {
        u.id: u for u in User.objects.filter(role="student").only("id", "roll_number")
    }
    for a in Attendance.objects.all().order_by("date", "id"):
        student = student_by_id.get(a.student_id)
        ws_att.append(
            [
                a.date.isoformat() if a.date else "",
                student.roll_number if student and student.roll_number else "",
                a.subject or "",
                a.status or "",
                a.attended_hours if a.attended_hours is not None else "",
                a.total_hours if a.total_hours is not None else "",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="attendance_data.xlsx"'
    wb.save(response)
    return response


def _parse_non_negative_int(value, allow_empty=False):
    """Parse value to non-negative int. Return (int or None, error_message or None)."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return (None, None) if allow_empty else (None, "Value is required.")
    try:
        n = int(float(str(value).strip()))
        if n < 0:
            return (None, "Must be a non-negative integer.")
        return (n, None)
    except (ValueError, TypeError):
        return (None, "Must be a non-negative integer.")


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_attendance_upload_view(request):
    """Upload attendance in bulk from an Excel (.xlsx) file.

    Required columns: roll_number, subject, date.
    Optional: attended_hours, total_hours, status.
    - When attended_hours and total_hours are both provided (valid numbers): status is optional and auto-calculated.
    - When attended_hours/total_hours are empty or missing: default is 1 hour (one session); status is required (Present/Absent).
    Duplicates (same student + subject + date) are skipped or updated.
    Only admin or faculty may call this endpoint.
    """
    if request.user.role not in ['admin', 'faculty'] and not request.user.is_superuser:
        return Response({"detail": "Only admin or faculty can upload attendance."}, status=403)

    upload = request.FILES.get('file')
    if not upload:
        return Response({"detail": "No file uploaded. Use form field 'file'."}, status=400)
    if not str(upload.name).lower().endswith('.xlsx'):
        return Response({"detail": "Invalid file type. Please upload an .xlsx file."}, status=400)

    try:
        file_content = upload.read()
        wb = load_workbook(filename=BytesIO(file_content), data_only=True)
    except Exception as e:
        return Response(
            {"detail": "Could not read Excel file. Make sure it is a valid .xlsx file.", "error": str(e)},
            status=400,
        )

    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Excel file is empty."}, status=400)

        header_row = rows[0]
        raw_headers = [str(v).strip() if v is not None else '' for v in header_row]
        # Normalize: accept "Roll Number", "roll number", "roll_number" etc. -> canonical names
        canonical = {
            'roll_number': ['roll_number', 'roll number', 'rollnumber'],
            'subject': ['subject'],
            'date': ['date'],
            'attended_hours': ['attended_hours', 'attended hours', 'attendedhours'],
            'total_hours': ['total_hours', 'total hours', 'totalhours'],
            'status': ['status'],
        }
        required_cols = ['roll_number', 'subject', 'date']
        header_to_canonical = {}
        for canon, variants in canonical.items():
            for v in variants:
                header_to_canonical[v.strip().lower().replace(' ', '_')] = canon
            header_to_canonical[canon] = canon
        # Build normalized list: each cell -> canonical name or ''
        headers_normalized = []
        for h in raw_headers:
            key = h.lower().replace(' ', '_').strip()
            headers_normalized.append(header_to_canonical.get(key, ''))
        headers = headers_normalized
        missing = [c for c in required_cols if c not in headers]
        if missing:
            return Response(
                {
                    "detail": "Missing required columns in header row.",
                    "missing_columns": missing,
                    "expected_columns": required_cols,
                    "received_headers": raw_headers,
                },
                status=400,
            )
        # When hours columns are absent, status is required (for default 1-hour rows)
        has_hours_cols = 'attended_hours' in headers and 'total_hours' in headers
        has_status_col = 'status' in headers
        if not has_hours_cols and not has_status_col:
            return Response(
                {
                    "detail": "When attended_hours and total_hours columns are not present, status column is required (Present/Absent).",
                    "received_headers": raw_headers,
                },
                status=400,
            )

        def _col_index(name):
            return headers.index(name) if name in headers else -1

        idx = {name: _col_index(name) for name in required_cols}
        if has_hours_cols:
            idx['attended_hours'] = headers.index('attended_hours')
            idx['total_hours'] = headers.index('total_hours')
        if has_status_col:
            idx['status'] = headers.index('status')

        # Preload students for faster lookup
        students_by_roll = {
            (u.roll_number or '').strip().upper(): u
            for u in User.objects.filter(role='student').exclude(roll_number__isnull=True)
        }

        created_count = 0
        updated_count = 0
        skipped_count = 0
        skipped_missing_student = 0
        skipped_missing_subject = 0
        skipped_invalid = 0
        error_rows = []

        def parse_date(value):
            if isinstance(value, datetime.date):
                return value if not isinstance(value, datetime.datetime) else value.date()
            if isinstance(value, datetime.datetime):
                return value.date()
            if value is None:
                return None
            if isinstance(value, (int, float)):
                try:
                    serial = int(round(value))
                    if serial < 1:
                        return None
                    return (datetime.datetime(1899, 12, 31) + datetime.timedelta(days=serial)).date()
                except (ValueError, OverflowError):
                    return None
            text = str(value).strip()
            if not text:
                return None
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
                try:
                    return datetime.datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
            return None

        for row_number, row in enumerate(rows[1:], start=2):
            if row is None:
                continue
            if all((cell is None or str(cell).strip() == '') for cell in row):
                continue

            def _get(col_name):
                i = idx.get(col_name, -1)
                if i < 0 or i >= len(row):
                    return ''
                value = row[i]
                return '' if value is None else str(value).strip()

            roll_number = _get('roll_number')
            subject_raw = _get('subject')
            date_raw = row[idx['date']] if idx['date'] < len(row) else None
            attended_raw = _get('attended_hours')
            total_raw = _get('total_hours')
            status_raw = _get('status')

            if not roll_number or not subject_raw:
                skipped_invalid += 1
                error_rows.append({"row": row_number, "reason": "Missing roll_number or subject."})
                continue

            student = students_by_roll.get(roll_number.strip().upper())
            if not student:
                skipped_missing_student += 1
                error_rows.append({"row": row_number, "reason": f"Student with roll_number '{roll_number}' not found."})
                continue

            subj = Subject.objects.filter(name__iexact=subject_raw).first() or Subject.objects.filter(code__iexact=subject_raw).first()
            if not subj:
                skipped_missing_subject += 1
                error_rows.append({"row": row_number, "reason": f"Subject '{subject_raw}' not found."})
                continue

            subject_value = subj.code or subj.name or subject_raw

            date_value = parse_date(date_raw)
            if not date_value:
                skipped_invalid += 1
                error_rows.append({"row": row_number, "reason": f"Invalid date value '{date_raw}'."})
                continue

            attended_hours_val, err_att = _parse_non_negative_int(attended_raw, allow_empty=True)
            total_hours_val, err_tot = _parse_non_negative_int(total_raw, allow_empty=True)
            hours_provided = attended_hours_val is not None and total_hours_val is not None and total_hours_val > 0

            if hours_provided:
                if err_att:
                    skipped_invalid += 1
                    error_rows.append({"row": row_number, "reason": f"attended_hours: {err_att}"})
                    continue
                if err_tot:
                    skipped_invalid += 1
                    error_rows.append({"row": row_number, "reason": f"total_hours: {err_tot}"})
                    continue
                if attended_hours_val > total_hours_val:
                    skipped_invalid += 1
                    error_rows.append({"row": row_number, "reason": "attended_hours cannot exceed total_hours."})
                    continue
                status_value = 'present' if attended_hours_val > 0 else 'absent'
            else:
                # Default 1 hour: require status (Present → 1 attended, Absent → 0)
                if not status_raw:
                    skipped_invalid += 1
                    error_rows.append({"row": row_number, "reason": "When attended_hours/total_hours are empty, status (Present/Absent) is required."})
                    continue
                status_lower = status_raw.strip().lower()
                if status_lower in ('present', 'p'):
                    attended_hours_val = 1
                    total_hours_val = 1
                    status_value = 'present'
                elif status_lower in ('absent', 'a'):
                    attended_hours_val = 0
                    total_hours_val = 1
                    status_value = 'absent'
                else:
                    skipped_invalid += 1
                    error_rows.append({"row": row_number, "reason": f"Invalid status '{status_raw}'. Use Present or Absent when hours are not given."})
                    continue

            att_obj, created = Attendance.objects.get_or_create(
                student=student,
                subject=subject_value,
                date=date_value,
                defaults={
                    "status": status_value,
                    "attended_hours": attended_hours_val,
                    "total_hours": total_hours_val,
                },
            )
            if created:
                created_count += 1
            else:
                updated_fields = []
                if att_obj.status != status_value:
                    att_obj.status = status_value
                    updated_fields.append("status")
                if att_obj.attended_hours != attended_hours_val:
                    att_obj.attended_hours = attended_hours_val
                    updated_fields.append("attended_hours")
                if att_obj.total_hours != total_hours_val:
                    att_obj.total_hours = total_hours_val
                    updated_fields.append("total_hours")
                if updated_fields:
                    att_obj.save(update_fields=updated_fields)
                    updated_count += 1
                else:
                    skipped_count += 1

        return Response(
            {
                "created": created_count,
                "updated": updated_count,
                "skipped": skipped_count,
                "skipped_missing_student": skipped_missing_student,
                "skipped_missing_subject": skipped_missing_subject,
                "skipped_invalid": skipped_invalid,
                "errors": error_rows,
            }
        )
    except Exception as e:
        return Response(
            {"detail": "Server error while processing upload.", "error": str(e)},
            status=500,
        )
